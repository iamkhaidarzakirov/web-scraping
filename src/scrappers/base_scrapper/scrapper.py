import logging
import os

from openpyxl.styles import Font, Border, Side
import pandas as pd

from helpers.captcha_helper.captcha_helper import CaptchaHelper
from helpers.proxies_helper.proxies_helper import ProxiesHelper
from helpers.pytesseract_helper.pytesseract_helper import PytesseractHelper
from config import settings


class BaseScrapper:
    def __init__(self) -> None:
        self.proxies_helper = ProxiesHelper()
        self.captcha_helper = CaptchaHelper()
        self.pytesseract_helper = PytesseractHelper()
        self.logger =  logging.getLogger("streamLogger")
    
    def add_rows_to_xlsx(self, file_name: str, source: dict, sheet_title: str = 'Sheet1') -> None:
        """
        Add one or many rows (depending on the DataFrame) to a .xlsx spreadsheet.

        This method allows for:
        - Adding rows to an existing sheet or creating a new sheet if it does not exist.
        - Handling dynamic table headers: it automatically adds new headers if they are missing.
        - Ensuring that the order of columns is preserved when adding data, even if new columns are added to the sheet.
        """
        
        if isinstance(source, list):
            df = pd.json_normalize(source)
        else:
            df = pd.json_normalize([source])

        path = os.path.join(settings.SHEETS_DIR, file_name)
        
        if not os.path.exists(path):
            self.logger.info(f"CREATE NEW EXCEL FILE ON: [{path}]")

            df.to_excel(path, sheet_name=sheet_title, index=False) # Create new excel file
        else:
            with pd.ExcelWriter(path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                new_headers = df.columns.to_list()
                # Work with existing excel file and worksheet
                if sheet_title in writer.sheets:
                    self.logger.info(f"WORK WITH EXISTING FILE AND SHEET: [{sheet_title}]")

                    sheet = writer.sheets[sheet_title]
                    old_headers = set([cell.value for cell in sheet[1]])
                    different_headers = list(set(new_headers).difference(old_headers))
                    
                    if different_headers:
                        self.logger.info(f"ADD NEW HEADERS: [{different_headers}]")
                        
                        end = sheet.max_column
                        bold_font = Font(bold=True)
                        thick_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        for i, header in enumerate(different_headers, start=1):
                            cell = sheet.cell(row=1, column=(end + i), value=header)
                            cell.font = bold_font
                            cell.border = thick_border
                        
                        new_headers = [cell.value for cell in sheet[1]] # Get updated headers ordering
                        df = df[new_headers] # Sync DF headers with existing updated sheet
                        
                        # writer.book.save(path)
                    
                    df.to_excel(writer, startrow=sheet.max_row, sheet_name=sheet_title, index=False, header=False)
                else:
                    self.logger.info(f"CREATE NEW SHEET IN EXISTING FILE: [{sheet_title}]")
                    df.to_excel(writer, startrow=0, sheet_name=sheet_title, index=False)
